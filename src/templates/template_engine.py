"""
Template Engine

Apply professional templates to Excel workbooks
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from typing import Optional
import pandas as pd


@dataclass
class TemplateConfig:
    """Configuration for Excel templates"""
    
    # Colors (hex format)
    primary_color: str = "1E3A8A"      # Navy blue
    secondary_color: str = "3B82F6"    # Lighter blue
    accent_color: str = "10B981"       # Green
    header_bg: str = "F3F4F6"          # Light gray
    
    # Fonts
    header_font: str = "Arial"
    body_font: str = "Arial"
    header_size: int = 12
    body_size: int = 10
    
    # Layout
    freeze_panes: bool = True
    auto_filter: bool = True
    column_width: str = "auto"  # 'auto' or specific number
    
    # Branding
    company_name: str = ""
    header_text: str = ""


class TemplateEngine:
    """Engine for applying templates to Excel workbooks"""
    
    def __init__(self, template_name: str = "professional", config: Optional[TemplateConfig] = None):
        """
        Initialize template engine
        
        Args:
            template_name: Name of predefined template  
            config: Custom template configuration
        """
        self.template_name = template_name
        self.config = config or self._get_predefined_config(template_name)
    
    def _get_predefined_config(self, name: str) -> TemplateConfig:
        """Get predefined template configuration"""
        templates = {
            "professional": TemplateConfig(
                primary_color="1E3A8A",
                secondary_color="3B82F6",
                header_bg="E5E7EB",
                company_name="Professional Template"
            ),
            "modern": TemplateConfig(
                primary_color="6366F1",
                secondary_color="8B5CF6",
                accent_color="EC4899",
                header_bg="F5F3FF",
                company_name="Modern Template"
            ),
            "financial": TemplateConfig(
                primary_color="065F46",
                secondary_color="10B981",
                accent_color="FBBF24",
                header_bg="ECFDF5",
                company_name="Financial Template"
            )
        }
        
        return templates.get(name, templates["professional"])
    
    def apply_to_worksheet(self, worksheet, has_header: bool = True):
        """
        Apply template styling to a worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            has_header: Whether first row is a header
        """
        # Style header row
        if has_header and worksheet.max_row > 0:
            self._style_header(worksheet)
        
        # Auto-adjust column widths
        if self.config.column_width == "auto":
            self._auto_size_columns(worksheet)
        
        # Apply freeze panes
        if self.config.freeze_panes and has_header:
            worksheet.freeze_panes = "A2"
        
        # Apply auto-filter
        if self.config.auto_filter and has_header:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Style data cells
        self._style_data_cells(worksheet, start_row=2 if has_header else 1)
    
    def _style_header(self, worksheet):
        """Apply styling to header row"""
        header_fill = PatternFill(
            start_color=self.config.header_bg,
            end_color=self.config.header_bg,
            fill_type="solid"
        )
        
        header_font = Font(
            name=self.config.header_font,
            size=self.config.header_size,
            bold=True,
            color=self.config.primary_color
        )
        
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False
        )
        
        border_side = Side(style="thin", color="D1D5DB")
        header_border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side
        )
        
        # Apply to all cells in first row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
    
    def _style_data_cells(self, worksheet, start_row: int = 2):
        """Apply styling to data cells"""
        body_font = Font(
            name=self.config.body_font,
            size=self.config.body_size
        )
        
        body_alignment = Alignment(
            horizontal="left",
            vertical="center"
        )
        
        # Light border
        border_side = Side(style="thin", color="E5E7EB")
        cell_border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side
        )
        
        # Alternate row colors
        light_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid"
        )
        
        alt_fill = PatternFill(
            start_color="F9FAFB",
            end_color="F9FAFB",
            fill_type="solid"
        )
        
        # Apply to all data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            # Alternate row fill
            fill = alt_fill if (row_idx - start_row) % 2 == 1 else light_fill
            
            for cell in row:
                if cell.value is not None:
                    cell.font = body_font
                    cell.alignment = body_alignment
                    cell.border = cell_border
                    cell.fill = fill
    
    def _auto_size_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max 50 chars
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_company_header(self, worksheet):
        """Add company header to worksheet"""
        if self.config.company_name or self.config.header_text:
            # Insert a new row at the top
            worksheet.insert_rows(1)
            
            # Merge cells for header
            worksheet.merge_cells('A1:' + get_column_letter(worksheet.max_column) + '1')
            
            # Set header text
            header_cell = worksheet['A1']
            header_cell.value = self.config.company_name or self.config.header_text
            
            # Style header
            header_cell.font = Font(
                name=self.config.header_font,
                size=14,
                bold=True,
                color=self.config.primary_color
            )
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            header_cell.fill = PatternFill(
                start_color=self.config.header_bg,
                end_color=self.config.header_bg,
                fill_type="solid"
            )
            
            # Adjust row height
            worksheet.row_dimensions[1].height = 25
